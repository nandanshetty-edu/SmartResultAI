from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

from services.student_service import StudentService
from services.result_service import ResultService
from services.mark_service import MarkService
from services.subject_service import SubjectService


class ExcelService:

    @staticmethod
    def load_template(excel_path):

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        return workbook, sheet

    @staticmethod
    def build_subject_map(sheet):

        subject_map = {}

        col = 5  # Column E

        while col <= sheet.max_column:

            code = sheet.cell(row=10, column=col).value

            if code:

                code = (
                    str(code)
                    .replace("\n", "")
                    .replace(" ", "")
                    .strip()
                    .upper()
                )

                if code == "RESULT":
                    col += 3
                    continue

                codes = code.split("/")

                for subject_code in codes:

                    subject_code = subject_code.strip()

                    subject_map[subject_code] = {
                        "internal": col,
                        "external": col + 1,
                        "total": col + 2,
                    }

            col += 3

        return subject_map

    @staticmethod
    def build_result_column_map(sheet):

        column_map = {}

        for col in range(1, sheet.max_column + 1):

            value = sheet.cell(row=10, column=col).value

            if not value:
                continue

            value = str(value).strip().upper()

            if value == "RESULT":
                column_map["result"] = col

            elif value == "TOTAL":
                column_map["grand_total"] = col

            elif value == "PERCENTAGE":
                column_map["percentage"] = col

        return column_map

    @staticmethod
    def write_summary(sheet, row, result, marks, result_columns):

        grand_total = sum(
            mark.total for mark in marks if mark.total is not None
        )

        max_marks = len(marks) * 100

        percentage = (
            round((grand_total / max_marks) * 100, 2)
            if max_marks > 0 else 0
        )

        if "result" in result_columns:
            sheet.cell(
                row=row,
                column=result_columns["result"]
            ).value = result.overall_result

        if "grand_total" in result_columns:
            sheet.cell(
                row=row,
                column=result_columns["grand_total"]
            ).value = grand_total

        if "percentage" in result_columns:
            sheet.cell(
                row=row,
                column=result_columns["percentage"]
            ).value = percentage

    @staticmethod
    def write_student_marks(
        sheet,
        row,
        exam,
        subject_map,
        result_columns
    ):

        usn = sheet.cell(row=row, column=2).value

        if not usn:
            return False

        usn = str(usn).strip()

        student = StudentService.find_by_usn(usn)

        if not student:
            return True

        result = ResultService.get_by_student_exam(
            student.id,
            exam.id
        )

        if not result:
            return True

        marks = MarkService.get_marks(result.id)

        for mark in marks:

            subject = SubjectService.get_by_id(mark.subject_id)

            if not subject:
                continue

            if subject.subject_code not in subject_map:
                continue

            columns = subject_map[subject.subject_code]

            sheet.cell(
                row=row,
                column=columns["internal"]
            ).value = mark.internal

            sheet.cell(
                row=row,
                column=columns["external"]
            ).value = mark.external

            sheet.cell(
                row=row,
                column=columns["total"]
            ).value = mark.total

        ExcelService.write_summary(
            sheet,
            row,
            result,
            marks,
            result_columns
        )

        return True

    @staticmethod
    def save_workbook(workbook, output_path):

        Path(output_path).parent.mkdir(
            parents=True,
            exist_ok=True
        )

        workbook.save(output_path)

        return output_path

    @staticmethod
    def generate(excel_path, exam):

        workbook, sheet = ExcelService.load_template(
            excel_path
        )

        subject_map = ExcelService.build_subject_map(
            sheet
        )

        result_columns = ExcelService.build_result_column_map(
            sheet
        )

        row = 12

        while True:

            usn = sheet.cell(
                row=row,
                column=2
            ).value

            if not usn:
                break

            ExcelService.write_student_marks(
                sheet,
                row,
                exam,
                subject_map,
                result_columns
            )

            row += 1

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        output_path = (
            f"outputs/result_{timestamp}.xlsx"
        )

        ExcelService.save_workbook(
            workbook,
            output_path
        )

        return output_path